from .helper import normalize
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
import os
from typing import Any


def get_distributor_info(acronym: str) -> dict[str, Any]:
    distributors_sheet = _load_distributors_sheet()
    
    name = _load_value(
        column_name='NOME',
        acronym=acronym,
        from_sheet=distributors_sheet
    )

    company_code = _load_value(
        column_name='CÓDIGO',
        acronym=acronym,
        from_sheet=distributors_sheet
    )

    agent_id = _load_value(
        column_name='ID AGENTE',
        acronym=acronym,
        from_sheet=distributors_sheet
    )

    concession_id = _load_value(
        column_name='ID CONCESSÃO',
        acronym=acronym,
        from_sheet=distributors_sheet
    )

    return {
        'Nome': name,
        'Código da Empresa': company_code,
        'ID Agente': agent_id,
        'ID Concessão': concession_id
    }


def _load_distributors_sheet() -> Worksheet:
    file_path = os.path.join(os.path.dirname(__file__), "../../distribuidoras.xlsx")
    file_path = os.path.abspath(file_path)

    workbook = load_workbook(file_path, keep_links=False, read_only=True, data_only=True)
    return workbook.active


def _load_value(column_name: str, acronym: str, from_sheet: Worksheet) -> Any:
    worksheet = from_sheet
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    header = list(header_row)

    try:
        aimed_column = header.index(column_name)
        acronym_column = header.index("SIGLA")
    except ValueError:
        raise ValueError(f"Column '{column_name}' or 'SIGLA' not found in header")
    
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        column_value = row[aimed_column]
        acronym_value = row[acronym_column]

        if normalize(acronym_value) ==  normalize(acronym):
            return column_value.strip() if isinstance(column_value, str) else column_value
    
    print(f"Sigla '{acronym}' não encontrada.")
    return None


def get_missing_acronyms() -> list[str]:
    codes = ['D14', 'D15', 'D21', 'D25', 'D32', 'D34', 'D38', 'D39', 'D43', 'D45', 'D47', 'D54', 'D56', 'D58', 'D61', 'D63', 'D64', 'D66']

    distribrutors_sheet = _load_distributors_sheet()
    header = [cell.value for cell in distribrutors_sheet[1]]
    
    try:
        acronym_index = header.index('SIGLA')
        code_index = header.index('CÓDIGO')
    except ValueError:
        raise Exception("As colunas 'SIGLA' e 'CÓDIGO' precisam estar presentes no cabeçalho")
    
    found_acronyms = []

    for row in distribrutors_sheet.iter_rows(min_row=2, values_only=True):
        code_value = row[code_index]
        acronym_value = row[acronym_index]

        if code_value in codes:
            found_acronyms.append(acronym_value)

    return found_acronyms