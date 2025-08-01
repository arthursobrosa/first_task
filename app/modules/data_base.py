from openpyxl import Workbook, load_workbook
from typing import Literal
import os
from .agent import Agent
from .distributor_data import get_distributor_info, get_missing_acronyms
from .cover_data import get_process_date
from .market_data import get_market_sheet
from .helper import get_suffix, remove_duplicate_rows, load_tab_data
from tqdm import tqdm


def process_workbooks(agent: Agent):
    # _create_db_folders(agent)
    # _process_db_files(agent)
    # _process_db_files(agent)
    foo()



def foo():
    base_path = os.path.join(os.path.dirname(__file__), "../../")
    base_path = os.path.abspath(base_path)

    distributors_path = os.path.join(base_path, "Distribuidoras")
    folders_path = os.path.join(distributors_path, "Permissionárias")
    distributor_path = os.path.join(folders_path, "CERILUZ")

    files_to_append = []

    for type in ["Reajuste", "Revisão"]:
        type_path = os.path.join(distributor_path, type)

        file_names = [
            name for name in os.listdir(type_path)
            if (name.endswith(".xlsx"))
            and not name.startswith("~$")
        ]

        for file in file_names:
            file_path = os.path.join(type_path, file)
            file_workbook = load_workbook(file_path, data_only=True)

            try:
                new_workbook = _filtered_workbook(
                    acronym="CERILUZ",
                    process_type=type,
                    workbook=file_workbook,
                    agent=Agent.Permissionaria
                )

                file_suffix = get_suffix(file)
                temp_path = file_path.replace(file_suffix, f"_temp{file_suffix}")
                new_workbook.save(temp_path)
                files_to_append.append(temp_path)
            except Exception as error:
                print(f"Could not filter {file_path} workbook: {str(error)}")

    if files_to_append:
        output_folder_path = os.path.join(distributor_path, "Banco de Dados")
        os.makedirs(output_folder_path, exist_ok=True)

        output_path = os.path.join(output_folder_path, "CERILUZ_BANCO.xlsx")

        _mix_db_files(
            files=files_to_append,
            output_name=output_path
        )

        print(f"\nBanco de dados consolidado em {output_path}")

        for temp_file in files_to_append:
            os.remove(temp_file)
                      

def _mix_db_files(files: list[str], output_name: str):
    max_row_per_sheet = 1048576

    output_workbook = Workbook()
    output_worksheet = output_workbook.active
    output_worksheet.title = "BANCO DE DADOS"

    current_sheet = output_worksheet
    current_row_count = 0
    sheet_index = 0
    header = None

    for index, file in enumerate(files):
        file_workbook = load_workbook(file, data_only=True)
        file_worksheet = file_workbook.active

        min_row = 1 if index == 0 else 2
        max_row = file_worksheet.max_row

        for row_idx, row in enumerate(file_worksheet.iter_rows(min_row=min_row, max_row=max_row, values_only=True), start=min_row):
            if index == 0 and row_idx == 1:
                header = row
                current_sheet.append(header)
                current_row_count += 1
                continue

            if current_row_count >= max_row_per_sheet:
                sheet_index += 1
                current_sheet = output_workbook.create_sheet(title=f"BANCO DE DADOS - Ext. {sheet_index}")
                current_sheet.append(header)
                current_row_count = 1

            current_sheet.append(row)
            current_row_count += 1

    output_workbook.save(output_name)


def _process_db_files(agent: Agent):
    base_path = os.path.join(os.path.dirname(__file__), "../../")
    base_path = os.path.abspath(base_path)

    distributors_path = os.path.join(base_path, "Distribuidoras")
    folders_path = os.path.join(distributors_path, agent.path)

    distributors = [
        name for name in os.listdir(folders_path)
        if os.path.isdir(os.path.join(folders_path, name))
    ]

    files = []

    for distributor in distributors:
        distributor_path = os.path.join(folders_path, distributor)
        db_path = os.path.join(distributor_path, "Banco de Dados")

        if not os.path.isdir(db_path):
            continue

        file_names = [
            name for name in os.listdir(db_path)
            if (name.endswith(".xlsx")) 
            and not name.startswith("~$")
        ]

        for file_name in file_names:
            file_path = os.path.join(db_path, file_name)
            files.append(file_path)

    if files:
        output_folder_path = os.path.join(distributors_path, "Banco de Dados")
        os.makedirs(output_folder_path, exist_ok=True)

        output_path = os.path.join(output_folder_path, f"BD-{agent.path}.xlsx")

        _mix_db_files(
            files=files,
            output_name=output_path
        )


def _create_db_folders(agent: Agent):
    base_path = os.path.join(os.path.dirname(__file__), "../../")
    base_path = os.path.abspath(base_path)

    distributors_path = os.path.join(base_path, "Distribuidoras")
    folders_path = os.path.join(distributors_path, agent.path)

    distributors = [
        name for name in os.listdir(folders_path)
        if os.path.isdir(os.path.join(folders_path, name))
    ]

    for distributor in tqdm(distributors, desc=f"Processando {agent.path}..."):
        distributor_path = os.path.join(folders_path, distributor)

        temp_file_paths = []

        if not os.path.isdir(distributor_path):
            print(f"\n{distributor_path} não é uma pasta")
            continue

        for type in ["Reajuste", "Revisão"]:
            type_path = os.path.join(distributor_path, type)

            if not os.path.isdir(type_path):
                print(f"\n{type_path} não é uma pasta")
                continue

            file_names = [
                name for name in os.listdir(type_path)
                if (name.endswith(".xlsx") or name.endswith(".xlsm")) 
                and not name.startswith("~$")
            ]

            for file_name in tqdm(file_names, desc=f"{distributor} - {type}", leave=False):
                file_path = os.path.join(type_path, file_name)
                file_workbook = load_workbook(file_path, data_only=True)

                try:
                    new_workbook = _filtered_workbook(
                        acronym=distributor,
                        process_type=type,
                        workbook=file_workbook,
                        agent=agent
                    )

                    file_suffix = get_suffix(file_name)
                    temp_path = file_path.replace(file_suffix, f"_temp{file_suffix}")
                    new_workbook.save(temp_path)
                    temp_file_paths.append(temp_path)
                except Exception as error:
                    print(f"\nFalha ao filtrar planilha em {file_path}: {str(error)}")


        if temp_file_paths:
            output_folder_path = os.path.join(distributor_path, "Banco de Dados")
            os.makedirs(output_folder_path, exist_ok=True)

            output_path = os.path.join(output_folder_path, f"{distributor}_BANCO.xlsx")

            _join_data_base_sheets(
                files=temp_file_paths,
                output_name=output_path
            )

            print(f"\nBanco de dados consolidado em {output_path}")

            for temp_file_path in temp_file_paths:
                os.remove(temp_file_path)


def _filtered_workbook(acronym: str, process_type: Literal['Reajuste', 'Revisão'], workbook: Workbook, agent: Agent) -> Workbook:
    distributor_info = get_distributor_info(acronym)

    process_date = get_process_date(
        workbook=workbook,
        agent=agent
    )

    distributor_info['Processo'] = process_type
    distributor_info['Data de Reajuste/Revisão em Processamento'] = process_date

    market_sheet = get_market_sheet(
        workbook=workbook,
        agent=agent
    )

    new_workbook = Workbook()
    new_tab = new_workbook.active
    new_tab.title = "BANCO DE DADOS"

    distributor_header = list(distributor_info.keys())
    market_header = [cell.value for cell in next(market_sheet.iter_rows(min_row=1, max_row=1))]

    new_tab.append(distributor_header + market_header)

    max_row = market_sheet.max_row
    max_col = market_sheet.max_column

    for row in market_sheet.iter_rows(min_row=2, max_row=max_row, max_col=max_col, values_only=True):
        if not all(cell is None for cell in row):
            new_row = list(distributor_info.values()) + list(row)
            new_tab.append(new_row)

    return new_workbook


def _join_data_base_sheets(files: list[str], output_name: str):
    all_data = []
    total_columns = []
    seen_columns = set()

    for file in tqdm(files, desc="Arquivos Processados"):
        data = load_tab_data(
            file_path=file,
            tab_name='BANCO DE DADOS'
        )

        all_data.extend(data)

        for row in data:
            for column in row.keys():
                if column not in seen_columns:
                    seen_columns.add(column)
                    total_columns.append(column)

    new_workbook = Workbook()
    tab = new_workbook.active
    tab.title = 'BANCO DE DADOS'
    tab.append(total_columns)

    print("Adicionando linhas à nova planilha...")
    for row in tqdm(all_data, desc="Linhas inseridas"):
        ordered_row = [row.get(column, None) for column in total_columns]
        tab.append(ordered_row)

    new_workbook.save(output_name)
    print(f"Arquivo conjunto salvo como '{output_name}'")


def _go_through_data_base_files(agent: Agent):
    files = []

    current_path = os.path.abspath(os.path.dirname(__file__))
    base_path = os.path.abspath(os.path.join(current_path, "../../"))

    distributors_path = os.path.join(base_path, "Distribuidoras")

    output_path = os.path.join(distributors_path, "Banco de Dados")
    os.makedirs(output_path, exist_ok=True)

    folders_path = os.path.join(distributors_path, agent.path)

    distributors = [name for name in os.listdir(folders_path)
                      if os.path.isdir(os.path.join(folders_path, name))]
    
    for distributor in tqdm(distributors, desc=f"Processando {agent.path}..."):
        distributor_path = os.path.join(folders_path, distributor)
        data_base_path = os.path.join(distributor_path, "Banco de Dados")

        if not os.path.isdir(data_base_path):
            continue

        for file_name in os.listdir(data_base_path):
            if (file_name.endswith(".xlsx") or file_name.endswith(".xlsm")) and not file_name.startswith("~$"):
                file_path = os.path.join(data_base_path, file_name)
                files.append(file_path)
                break

    file_output_path = os.path.join(output_path, f"BD-{agent.path}.xlsx")

    _join_data_base_sheets(
        files=files,
        output_name=file_output_path
    )

    print(f"\nBanco de dados final salvo em: {file_output_path}")