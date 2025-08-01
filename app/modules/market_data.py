from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from .agent import Agent
from .helper import remove_initial_blank_rows


def get_market_sheet(workbook: Workbook, agent: Agent) -> Worksheet:
    expected_columns = [
        "Código", "TipoMercado", "Modalidade", "Subgrupo", "Classe",
        "Subclasse", "Detalhe", "Agente", "Posto", "OPÇÃO", "AnoMes",
        "D", "Daj", "TUSD_E", "TUSD_Eaj", "TE_E", "TE_Eaj"
    ]

    market_worksheet = workbook[agent.market_tab_name]
    remove_initial_blank_rows(market_worksheet)
    header = [cell.value for cell in next(
        market_worksheet.iter_rows(min_row=1, max_row=1))]

    aimed_indexes = {}

    for column in expected_columns:
        if column in header:
            aimed_indexes[column] = header.index(column)
        elif column == 'Código' and 'id_MercProj' in header:
            aimed_indexes[column] = header.index('id_MercProj')

    output_workbook = Workbook()
    new_tab = output_workbook.active
    new_tab.title = 'Mercado_Formatado'
    new_tab.append(list(aimed_indexes.keys()))

    for row in market_worksheet.iter_rows(min_row=2, values_only=True):
        new_row = [row[idx] for idx in aimed_indexes.values()]
        new_tab.append(new_row)

    return new_tab
