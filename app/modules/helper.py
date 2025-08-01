from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from pathlib import Path
from typing import Optional
import unicodedata


def load_tab_data(file_path: str, tab_name: str) -> list[dict[str, any]]:
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    tab = workbook[tab_name]

    headers = [cell.value for cell in next(tab.iter_rows(min_row=1, max_row=1))]
    data = []

    for row in tab.iter_rows(min_row=2, values_only=True):
        if any(field is not None for field in row):
            row_dict = {header: value for header, value in zip(headers, row)}
            data.append(row_dict)

    return data


def remove_initial_blank_rows(sheet: Worksheet):
    while True:
        first_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]

        if all(cell is None or str(cell).strip() == '' for cell in first_row):
            sheet.delete_rows(1)
        else:
            break


def remove_duplicate_rows(file_path: str, tab_name: Optional[str] = None):
    original_workbook = load_workbook(file_path, keep_links=False, data_only=True)
    original_tab = original_workbook[tab_name] if tab_name else original_workbook.active

    seen_rows = set()
    unique_rows = []

    for row in original_tab.iter_rows(values_only=True):
        if row not in seen_rows:
            seen_rows.add(row)
            unique_rows.append(row)

    new_workbook = Workbook()
    new_tab = new_workbook.active
    new_tab.title = original_tab.title

    for row in unique_rows:
        new_tab.append(row)

    new_workbook.save(file_path)


def get_suffix(file_name: str) -> str:
    path = Path(file_name)
    return path.suffix


def normalize(text: str) -> str:
    if not isinstance(text, str):
        return ""
    
    text.strip().upper()

    return ''.join(
        c for c in unicodedata.normalize('NFD', text)
        if unicodedata.category(c) != 'Mn'
    )