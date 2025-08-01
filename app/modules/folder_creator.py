from openpyxl import load_workbook, Workbook
import os
from .agent import Agent
from tqdm import tqdm
from .distributor_data import get_distributor_info


# def identify_invalid_files(agent: Agent):
#     current_path = os.path.abspath(os.path.dirname(__file__))
#     base_path = os.path.abspath(os.path.join(current_path, "../../"))

#     distributors_path = os.path.join(base_path, "Distribuidoras")
#     folders_path = os.path.join(distributors_path, agent.path)

#     distributors = [name for name in os.listdir(folders_path)
#                       if os.path.isdir(os.path.join(folders_path, name))]

#     for distributor in tqdm(distributors, desc=f"Processando {agent.path}"):
#         distributor_path = os.path.join(folders_path, distributor)

#         for type in ['Reajuste', 'Revisão']:
#             type_path = os.path.join(distributor_path, type)

#             file_names = [
#                 name for name in os.listdir(type_path)
#                 if name.endswith(".xls") and not name.startswith("~$")
#             ]

#             print(file_names)


# def _analyze_misplaced_files():
#     file_path = "/Users/arthursobrosa/Desktop/Tarifa1/Distribuidoras/Permissionárias/CERPRO/Reajuste/PERSAS_CERPRO_2015.xlsx"

#     file_workbook = load_workbook(file_path, read_only=True, data_only=True)

#     file_concession_id = _get_concession_id(
#         workbook=file_workbook,
#         agent=Agent.Permissionaria
#     )

#     print(file_concession_id)


# def _analyze_misplaced_files(acronym: str, agent: Agent):
#     current_path = os.path.abspath(os.path.dirname(__file__))
#     base_path = os.path.abspath(os.path.join(current_path, "../../"))

#     distributors_path = os.path.join(base_path, "Distribuidoras")
#     folders_path = os.path.join(distributors_path, agent.path)

#     distributor_path = os.path.join(folders_path, acronym)
#     distributor_info = get_distributor_info(acronym)
#     distributor_concession_id = distributor_info["ID Concessão"]

#     for type in ['Reajuste', 'Revisão']:
#         type_path = os.path.join(distributor_path, type)

#         file_names = [
#             name for name in os.listdir(type_path)
#             if (name.endswith(".xlsx") or name.endswith(".xlsm"))
#             and not name.startswith("~$")
#         ]

#         for file_name in tqdm(file_names, desc=f"{acronym} - {type}", leave=False):
#             file_path = os.path.join(type_path, file_name)
#             file_workbook = load_workbook(file_path, read_only=True, data_only=True)

#             file_concession_id = _get_concession_id(
#                 workbook=file_workbook,
#                 agent=agent
#             )

#             if file_concession_id != distributor_concession_id:
#                 print(f"{file_path} - {distributor_concession_id}/{file_concession_id}")


def _analyze_misplaced_files(agent: Agent):
    current_path = os.path.abspath(os.path.dirname(__file__))
    base_path = os.path.abspath(os.path.join(current_path, "../../"))

    distributors_path = os.path.join(base_path, "Distribuidoras")
    folders_path = os.path.join(distributors_path, agent.path)

    distributors = [name for name in os.listdir(folders_path)
                      if os.path.isdir(os.path.join(folders_path, name))]

    for distributor in tqdm(distributors, desc=f"Processando {agent.path}"):
        distributor_path = os.path.join(folders_path, distributor)
        distributor_info = get_distributor_info(distributor)
        distributor_concession_id = distributor_info["ID Concessão"]

        for type in ['Reajuste', 'Revisão']:
            type_path = os.path.join(distributor_path, type)

            file_names = [
                name for name in os.listdir(type_path)
                if (name.endswith(".xlsx") or name.endswith(".xlsm"))
                and not name.startswith("~$")
            ]

            for file_name in tqdm(file_names, desc=f"{distributor} - {type}", leave=False):
                file_path = os.path.join(type_path, file_name)
                file_workbook = load_workbook(file_path, read_only=True, data_only=True)

                file_concession_id = _get_concession_id(
                    workbook=file_workbook,
                    agent=agent
                )

                if file_concession_id != distributor_concession_id:
                    print(f"{file_path} - {distributor_concession_id}/{file_concession_id}")


def _get_concession_id(workbook: Workbook, agent: Agent) -> int:
    cover = workbook['CAPA']

    if agent.concession_id_dn:
        defined_name = workbook.defined_names.get(agent.concession_id_dn)
        # defined_name = workbook.defined_names[agent.concession_id_dn]

        if defined_name:
            for tab_name, cell_ref in defined_name.destinations:
                tab_origin = workbook[tab_name]
                return tab_origin[cell_ref].value
        
    for coord in agent.concession_ids_coord:
        value = cover[coord].value

        if value and isinstance(value, int):
            return value


def create_folders_for(agent: Agent):
    current_path = os.path.abspath(os.path.dirname(__file__))
    base_path = os.path.abspath(os.path.join(current_path, "../../"))

    distributors_path = os.path.join(base_path, "Distribuidoras")
    folders_path = os.path.join(distributors_path, agent.path)
    os.makedirs(folders_path, exist_ok=True)

    distributors = _get_acronyms_from(agent)

    for distributor in distributors:
        distributor_path = os.path.join(folders_path, distributor)
        os.makedirs(distributor_path, exist_ok=True)

        for type in ["Reajuste", "Revisão"]:
            type_path = os.path.join(distributor_path, type)
            os.makedirs(type_path, exist_ok=True)


def _get_acronyms_from(agent: Agent) -> str:
    file_path = os.path.join(os.path.dirname(__file__), "../../distribuidoras.xlsx")
    file_path = os.path.abspath(file_path)

    workbook = load_workbook(file_path, data_only=True)
    worksheet = workbook.active

    header = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    acronym_index = header.index("SIGLA")
    agent_index = header.index("AGENTE")

    distributors_acronyms = []

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        acronym_cell = row[acronym_index]
        agent_cell = row[agent_index]

        if agent_cell == agent.type_name:
            distributors_acronyms.append(acronym_cell)

    return distributors_acronyms